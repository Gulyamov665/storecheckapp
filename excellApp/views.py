from datetime import date
from itertools import count

from django.http import HttpResponse
from openpyxl.styles import Font, Alignment
import openpyxl
from store.models import Visit, Sku, Details


def export_xlsx(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="visit_sku.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active

    # Sku names as column headers
    ws.cell(row=1, column=1, value="Territory")
    ws.cell(row=1, column=2, value="Trade")
    sku_names = list(Sku.objects.filter(user=request.user).values_list('sku_name', flat=True))

    for i, sku_name in enumerate(sku_names):
        column = i + 3  # Start at cell C1
        ws.cell(row=1, column=column, value=sku_name).alignment = Alignment(textRotation=90)

    sku_column_count = len(sku_names)

    details_names = list(Details.objects.filter().values_list('name', flat=True))
    for i, detail in enumerate(details_names):
        column_detail = i + 3
        detail_column = sku_column_count + column_detail
        ws.cell(row=1, column=detail_column, value=detail)

    coment = len(sku_names)+len(details_names)

    ws.cell(row=1, column=coment + 3, value='Comments')
    ws.cell(row=1, column=coment + 4, value='Percent')


    # Get all visits for the current user
    default_date = date.today()
    user_visits = Visit.objects.filter(visit_date=default_date).filter(user=request.user).prefetch_related('sku')

    # Visit and Sku data
    for i, visit in enumerate(user_visits):
        row = i + 2  # Start at row 2
        ws.cell(row=row, column=1, value=visit.territory.territory_name)
        ws.cell(row=row, column=2, value=visit.trade.trade_name)

        ws.cell(row=row, column=coment + 3, value=visit.comment)

        # Get all Sku names
        all_sku_names = list(Sku.objects.filter(user=request.user).values_list('sku_name', flat=True))
        all_detail_names = Details.objects.filter(user=request.user)

        num_sku = 0
        num_sku_found = 0
        num_sku_not_found = 0

        for j, sku_name in enumerate(all_sku_names):
            column = j + 3

            if sku_name in [sku.sku_name for sku in visit.sku.all()]:
                ws.cell(row=row, column=column, value=1).font = Font(color="00ff00")
                num_sku_found += 1
            else:
                ws.cell(row=row, column=column, value=0).font = Font(color="FF0000")
                num_sku_not_found += 1
            num_sku += 1

        for g, detail in enumerate(all_detail_names):
            column_detail = g + 3
            detail_column = sku_column_count + column_detail

            if detail.name in [detail.name for detail in visit.detail.all()]:
                ws.cell(row=row, column=detail_column, value='+')
                
            else:
                ws.cell(row=row, column=detail_column, value=0)
                
        percentage_sku_found = round((num_sku_found / num_sku) * 100)
        ws.cell(row=row, column=coment + 4, value=f'{percentage_sku_found}%')


    wb.save(response)
    return response
