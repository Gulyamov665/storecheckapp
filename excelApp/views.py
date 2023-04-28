from datetime import date
import openpyxl
from django.http import HttpResponse
from openpyxl.descriptors import DateTime
from openpyxl.styles import Alignment, Font
from store.models import *


def export_xlsx(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="visit_sku.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active

    # Sku names as column headers
    ws.cell(row=1, column=1, value="Territory")
    ws.cell(row=1, column=2, value="Trade")

    sku_names = Sku.objects.filter(user=request.user).values_list('sku_name', flat=True)
    for i, sku_name in enumerate(sku_names):
        column = i + 3  # Start cell C1
        ws.cell(row=1, column=column, value=sku_name).alignment = Alignment(
            textRotation=90)
    column_sku = len(sku_names)

    name_united = United.objects.filter(user=request.user).values_list('name_detail', flat=True)
    for c, name_detail in enumerate(name_united):
        column_united = c + 3
        column_count = column_sku + column_united
        ws.cell(row=1, column=column_count, value=name_detail).alignment = Alignment(
            textRotation=90)
    column_w = len(name_united)+ len(sku_names)
    details_name = Details.objects.filter(user=request.user).values_list('name', flat=True)
    for d, name in enumerate(details_name):
        column_detail = d + 3
        column_all = column_w + column_detail
        ws.cell(row=1, column=column_all, value=name)
    column_ws = len(name_united) + len(sku_names)+len(details_name)
    ws.cell(row=1, column=column_ws+3, value="Percent")
    ws.cell(row=1, column=column_ws+4, value="Comment")
    # Get all visits for the current user
    default_date = date.today()
    user_visits = Visit.objects.filter(date=default_date).filter(
        user=request.user).prefetch_related('sku')

    # Visit and Sku data
    count = len(user_visits)
    total_counts = {}
    for i, visit in enumerate(user_visits):
        row = i + 2  # Start row 2
        ws.cell(row=row, column=2, value=visit.trade.name_trade)
        ws.cell(row=row, column=1, value=visit.territory.territory_name)
        ws.cell(row=row, column=column_ws+4, value=visit.comment)

        # Get all Sku names
        all_sku = Sku.objects.filter(user=request.user)

        num_sku = 0
        num_sku_found = 0

        for j, sku in enumerate(all_sku):
            column = j + 3

            if sku.sku_name in [sku.sku_name for sku in visit.sku.all()]:
                ws.cell(row=row, column=column,
                        value=1).font = Font(color="00ff00")
                num_sku_found += 1

            else:
                ws.cell(row=row, column=column,
                        value=0).font = Font(color="FF0000")

            if sku.id not in total_counts:
                total_counts[sku] = 0

            total_counts[sku] += 1
            num_sku += 1

        all_detail = Details.objects.filter(user=request.user)
        for l, detail in enumerate(all_detail):
            column_detail = l + 3
            column_all = column_w + column_detail

            if detail.name in [detail.name for detail in visit.detail.all()]:
                ws.cell(row=row, column=column_all, value=1)
            else:
                ws.cell(row=row, column=column_all, value=0)

        percent_sku_found = round((num_sku_found / num_sku) * 100, 2)
        ws.cell(row=row, column=column_ws+3 , value=f'{percent_sku_found}%')

    wb.save(response)

    return response
